import os
import shutil
import tempfile

from datetime import datetime

from openpyxl import load_workbook


# ============================================================
# CONFIGURACIÓN
# ============================================================

carpeta_programa = os.path.dirname(
    os.path.abspath(__file__)
)


archivo_excel = os.path.join(
    carpeta_programa,
    "productos_corregido.xlsx"
)


carpeta_backups = os.path.join(
    carpeta_programa,
    "backups"
)


os.makedirs(
    carpeta_backups,
    exist_ok=True
)


nombre_hoja_movimientos = "Movimientos"

stock_minimo = 5


# ============================================================
# CARGAR EXCEL
# ============================================================

def cargar_excel():

    if not os.path.isfile(
        archivo_excel
    ):

        raise Exception(
            f"No existe el archivo Excel:\n"
            f"{archivo_excel}"
        )

    try:

        return load_workbook(
            archivo_excel
        )

    except Exception as error:

        raise Exception(
            f"No se pudo abrir el Excel:\n"
            f"{error}"
        )


# ============================================================
# BUSCAR HOJA DE PRODUCTOS
# ============================================================

def buscar_hoja_productos(libro):

    for hoja in libro.worksheets:

        for fila in range(
            1,
            hoja.max_row + 1
        ):

            valores = []

            for columna in range(
                1,
                hoja.max_column + 1
            ):

                valor = hoja.cell(
                    row=fila,
                    column=columna
                ).value

                if valor is not None:

                    valores.append(
                        str(valor).strip()
                    )

            if (
                "Comidas" in valores
                and
                "Código de serie" in valores
            ):

                return hoja

    return None


# ============================================================
# BUSCAR ENCABEZADOS
# ============================================================

def buscar_encabezados(hoja):

    for fila in range(
        1,
        hoja.max_row + 1
    ):

        valores = []

        for columna in range(
            1,
            hoja.max_column + 1
        ):

            valor = hoja.cell(
                row=fila,
                column=columna
            ).value

            if valor is not None:

                valores.append(
                    str(valor).strip()
                )

        if (
            "Comidas" in valores
            and
            "Código de serie" in valores
        ):

            return fila

    return None


# ============================================================
# CONFIGURACIÓN DEL EXCEL
# ============================================================

def obtener_configuracion(libro):

    hoja = buscar_hoja_productos(
        libro
    )

    if hoja is None:

        raise Exception(
            "No se encontró la hoja de productos."
        )

    fila_encabezados = (
        buscar_encabezados(
            hoja
        )
    )

    if fila_encabezados is None:

        raise Exception(
            "No se encontró la fila de encabezados."
        )

    columna_producto = None
    columna_precio = None
    columna_cantidad = None
    columna_codigo = None

    for columna in range(
        1,
        hoja.max_column + 1
    ):

        valor = hoja.cell(
            row=fila_encabezados,
            column=columna
        ).value

        if valor is None:

            continue

        nombre = str(
            valor
        ).strip()

        if nombre == "Comidas":

            columna_producto = columna

        elif nombre == "valor":

            columna_precio = columna

        elif nombre == "Cantidad":

            columna_cantidad = columna

        elif nombre == "Código de serie":

            columna_codigo = columna

    if columna_producto is None:

        raise Exception(
            "No existe la columna 'Comidas'."
        )

    if columna_cantidad is None:

        raise Exception(
            "No existe la columna 'Cantidad'."
        )

    if columna_codigo is None:

        raise Exception(
            "No existe la columna 'Código de serie'."
        )

    return (
        hoja,
        fila_encabezados,
        columna_producto,
        columna_precio,
        columna_cantidad,
        columna_codigo
    )


# ============================================================
# CONVERTIR CANTIDAD
# ============================================================

def convertir_cantidad(valor):

    if valor is None:

        return 0

    try:

        return int(
            float(valor)
        )

    except (
        ValueError,
        TypeError
    ):

        return 0


# ============================================================
# BUSCAR PRODUCTO
# ============================================================

def buscar_producto_en_excel(
    hoja,
    fila_encabezados,
    columna_codigo,
    codigo
):

    codigo = str(
        codigo
    ).strip()

    for fila in range(
        fila_encabezados + 1,
        hoja.max_row + 1
    ):

        valor_codigo = hoja.cell(
            row=fila,
            column=columna_codigo
        ).value

        if valor_codigo is None:

            continue

        if (
            str(valor_codigo).strip()
            == codigo
        ):

            return fila

    return None


# ============================================================
# HOJA MOVIMIENTOS
# ============================================================

def obtener_hoja_movimientos(libro):

    if (
        nombre_hoja_movimientos
        in libro.sheetnames
    ):

        hoja = libro[
            nombre_hoja_movimientos
        ]

    else:

        hoja = libro.create_sheet(
            nombre_hoja_movimientos
        )

    encabezados = [
        "Fecha",
        "Código",
        "Producto",
        "Operación",
        "Cantidad",
        "Stock anterior",
        "Stock resultante"
    ]

    for columna, nombre in enumerate(
        encabezados,
        start=1
    ):

        if (
            hoja.cell(
                row=1,
                column=columna
            ).value is None
        ):

            hoja.cell(
                row=1,
                column=columna
            ).value = nombre

    return hoja


# ============================================================
# BACKUP
# ============================================================

def crear_backup():

    fecha = datetime.now().strftime(
        "%Y-%m-%d_%H-%M-%S"
    )

    nombre = (
        f"productos_{fecha}.xlsx"
    )

    ruta = os.path.join(
        carpeta_backups,
        nombre
    )

    shutil.copy2(
        archivo_excel,
        ruta
    )

    return ruta


# ============================================================
# GUARDAR EXCEL
# ============================================================

def guardar_excel(libro):

    ruta_temporal = None

    try:

        archivo_temporal = (
            tempfile.NamedTemporaryFile(
                prefix="stock_temporal_",
                suffix=".xlsx",
                dir=carpeta_programa,
                delete=False
            )
        )

        ruta_temporal = (
            archivo_temporal.name
        )

        archivo_temporal.close()

        libro.save(
            ruta_temporal
        )

        prueba = load_workbook(
            ruta_temporal,
            read_only=True
        )

        prueba.close()

        os.replace(
            ruta_temporal,
            archivo_excel
        )

        ruta_temporal = None

        return True

    except Exception as error:

        print(
            "ERROR AL GUARDAR EXCEL:",
            error
        )

        return False

    finally:

        if (
            ruta_temporal
            and
            os.path.exists(
                ruta_temporal
            )
        ):

            try:

                os.remove(
                    ruta_temporal
                )

            except Exception:

                pass


# ============================================================
# OBTENER PRODUCTOS
# ============================================================

def obtener_productos():

    libro = cargar_excel()

    try:

        (
            hoja,
            fila_encabezados,
            columna_producto,
            columna_precio,
            columna_cantidad,
            columna_codigo
        ) = obtener_configuracion(
            libro
        )

        productos = []

        for fila in range(
            fila_encabezados + 1,
            hoja.max_row + 1
        ):

            producto = hoja.cell(
                row=fila,
                column=columna_producto
            ).value

            codigo = hoja.cell(
                row=fila,
                column=columna_codigo
            ).value

            if (
                producto is None
                or
                codigo is None
            ):

                continue

            stock_actual = convertir_cantidad(
                hoja.cell(
                    row=fila,
                    column=columna_cantidad
                ).value
            )

            precio = None

            if columna_precio is not None:

                precio = hoja.cell(
                    row=fila,
                    column=columna_precio
                ).value

            productos.append({

                "codigo": str(
                    codigo
                ),

                "producto": str(
                    producto
                ),

                "precio": precio,

                "stock": stock_actual,

                "stock_bajo": (
                    stock_actual
                    <= stock_minimo
                )

            })

        return productos

    finally:

        libro.close()


# ============================================================
# BUSCAR PRODUCTO
# ============================================================

def buscar_producto(codigo):

    codigo = str(
        codigo
    ).strip()

    productos = obtener_productos()

    for producto in productos:

        if producto["codigo"] == codigo:

            return producto

    return None


# ============================================================
# REGISTRAR MOVIMIENTO
# ============================================================

def registrar_movimiento(
    codigo,
    operacion,
    cantidad
):

    codigo = str(
        codigo
    ).strip()

    operacion = str(
        operacion
    ).strip().upper()

    try:

        cantidad = int(
            cantidad
        )

    except (
        ValueError,
        TypeError
    ):

        return {
            "ok": False,
            "mensaje":
                "La cantidad debe ser un número entero."
        }

    if cantidad <= 0:

        return {
            "ok": False,
            "mensaje":
                "La cantidad debe ser mayor que 0."
        }

    if operacion not in (
        "ENTRADA",
        "SALIDA"
    ):

        return {
            "ok": False,
            "mensaje":
                "Operación inválida."
        }

    libro = cargar_excel()

    try:

        (
            hoja,
            fila_encabezados,
            columna_producto,
            columna_precio,
            columna_cantidad,
            columna_codigo
        ) = obtener_configuracion(
            libro
        )

        fila = buscar_producto_en_excel(
            hoja,
            fila_encabezados,
            columna_codigo,
            codigo
        )

        if fila is None:

            return {
                "ok": False,
                "mensaje":
                    f"Producto no encontrado: {codigo}"
            }

        producto = hoja.cell(
            row=fila,
            column=columna_producto
        ).value

        celda_stock = hoja.cell(
            row=fila,
            column=columna_cantidad
        )

        stock_anterior = convertir_cantidad(
            celda_stock.value
        )

        if operacion == "ENTRADA":

            stock_nuevo = (
                stock_anterior
                + cantidad
            )

        else:

            if cantidad > stock_anterior:

                return {
                    "ok": False,
                    "mensaje": (
                        "Stock insuficiente. "
                        f"Disponible: "
                        f"{stock_anterior}"
                    )
                }

            stock_nuevo = (
                stock_anterior
                - cantidad
            )

        try:

            ruta_backup = crear_backup()

        except Exception as error:

            return {
                "ok": False,
                "mensaje": (
                    "No se pudo crear el backup: "
                    f"{error}"
                )
            }

        celda_stock.value = (
            stock_nuevo
        )

        hoja_movimientos = (
            obtener_hoja_movimientos(
                libro
            )
        )

        nueva_fila = (
            hoja_movimientos.max_row + 1
        )

        datos = [
            datetime.now(),
            codigo,
            producto,
            operacion,
            cantidad,
            stock_anterior,
            stock_nuevo
        ]

        for columna, valor in enumerate(
            datos,
            start=1
        ):

            hoja_movimientos.cell(
                row=nueva_fila,
                column=columna
            ).value = valor

        if not guardar_excel(
            libro
        ):

            return {
                "ok": False,
                "mensaje":
                    "No se pudo guardar el Excel."
            }

        return {

            "ok": True,

            "mensaje":
                "Stock actualizado correctamente.",

            "producto":
                str(producto),

            "codigo":
                codigo,

            "operacion":
                operacion,

            "cantidad":
                cantidad,

            "stock_anterior":
                stock_anterior,

            "stock_nuevo":
                stock_nuevo,

            "backup":
                ruta_backup
        }

    finally:

        libro.close()


# ============================================================
# STOCK BAJO
# ============================================================

def obtener_stock_bajo():

    productos = obtener_productos()

    return [
        producto
        for producto in productos
        if producto["stock"]
        <= stock_minimo
    ]


# ============================================================
# MOVIMIENTOS
# ============================================================

def obtener_movimientos():

    libro = cargar_excel()

    try:

        hoja = obtener_hoja_movimientos(
            libro
        )

        movimientos = []

        for fila in range(
            2,
            hoja.max_row + 1
        ):

            codigo = hoja.cell(
                row=fila,
                column=2
            ).value

            if codigo is None:

                continue

            movimientos.append({

                "fecha":
                    hoja.cell(
                        row=fila,
                        column=1
                    ).value,

                "codigo":
                    str(codigo),

                "producto":
                    hoja.cell(
                        row=fila,
                        column=3
                    ).value,

                "operacion":
                    hoja.cell(
                        row=fila,
                        column=4
                    ).value,

                "cantidad":
                    hoja.cell(
                        row=fila,
                        column=5
                    ).value,

                "stock_anterior":
                    hoja.cell(
                        row=fila,
                        column=6
                    ).value,

                "stock_resultante":
                    hoja.cell(
                        row=fila,
                        column=7
                    ).value
            })

        return movimientos

    finally:

        libro.close()