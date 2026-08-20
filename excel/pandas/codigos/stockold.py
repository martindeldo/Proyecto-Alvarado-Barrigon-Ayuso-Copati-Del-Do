import os
import shutil
import tempfile
import traceback
from datetime import datetime

from openpyxl import load_workbook


# ============================================================
# CONFIGURACIÓN
# ============================================================

# Carpeta donde está este archivo stock.py
carpeta_programa = os.path.dirname(
    os.path.abspath(__file__)
)


# ============================================================
# ARCHIVO EXCEL PRINCIPAL
# ============================================================

archivo_excel = os.path.join(
    carpeta_programa,
    "productos_corregido.xlsx"
)


# ============================================================
# CARPETA DE BACKUPS
# ============================================================

carpeta_backups = os.path.join(
    carpeta_programa,
    "backups"
)

os.makedirs(
    carpeta_backups,
    exist_ok=True
)


# ============================================================
# CONFIGURACIÓN DEL SISTEMA
# ============================================================

nombre_hoja_movimientos = "Movimientos"

stock_minimo = 5


# ============================================================
# COMPROBAR QUE EL EXCEL EXISTE
# ============================================================

if not os.path.isfile(archivo_excel):

    raise Exception(
        "\nNo existe el archivo Excel:\n"
        f"{archivo_excel}\n\n"
        "Comprobá que el nombre del archivo sea correcto."
    )


# ============================================================
# ABRIR EXCEL
# ============================================================

try:

    libro = load_workbook(
        archivo_excel
    )

except Exception as error:

    raise Exception(
        "\nNo se pudo abrir el archivo Excel.\n\n"
        f"Archivo:\n{archivo_excel}\n\n"
        f"Error:\n{error}\n\n"
        "El archivo puede estar corrupto o no ser "
        "un archivo .xlsx válido."
    )


# ============================================================
# MOSTRAR HOJAS ENCONTRADAS
# ============================================================

print()
print(
    "=========================================="
)

print(
    "          HOJAS DEL EXCEL"
)

print(
    "=========================================="
)

for nombre in libro.sheetnames:

    print(
        f"- {nombre}"
    )

print()


# ============================================================
# BUSCAR HOJA DE PRODUCTOS
# ============================================================

def buscar_hoja_productos(libro):

    for hoja_busqueda in libro.worksheets:

        for fila in range(
            1,
            hoja_busqueda.max_row + 1
        ):

            valores = []

            for columna in range(
                1,
                hoja_busqueda.max_column + 1
            ):

                valor = hoja_busqueda.cell(
                    row=fila,
                    column=columna
                ).value

                if valor is not None:

                    valores.append(
                        str(valor).strip()
                    )

            if (
                "Comidas" in valores
                and "Código de serie" in valores
            ):

                return hoja_busqueda

    return None


hoja = buscar_hoja_productos(
    libro
)


if hoja is None:

    raise Exception(
        "\nNo se pudo encontrar la hoja de productos.\n\n"
        "El programa buscó una hoja que contenga las columnas:\n"
        "- Comidas\n"
        "- Código de serie\n\n"
        "Hojas encontradas:\n"
        + "\n".join(
            f"- {nombre}"
            for nombre in libro.sheetnames
        )
    )


print(
    f"Hoja de productos: {hoja.title}"
)


# ============================================================
# BUSCAR FILA DE ENCABEZADOS
# ============================================================

def buscar_encabezados(hoja):

    for fila in range(
        1,
        hoja.max_row + 1
    ):

        valores = []

        for columna in range(
            1,
            hoja.max_column + 1
        ):

            valor = hoja.cell(
                row=fila,
                column=columna
            ).value

            if valor is not None:

                valores.append(
                    str(valor).strip()
                )

        if (
            "Comidas" in valores
            and "Código de serie" in valores
        ):

            return fila

    return None


fila_encabezados = buscar_encabezados(
    hoja
)


if fila_encabezados is None:

    raise Exception(
        "\nNo se encontró la fila de encabezados."
    )


print(
    f"Fila de encabezados: {fila_encabezados}"
)


# ============================================================
# BUSCAR COLUMNAS
# ============================================================

columna_producto = None
columna_precio = None
columna_cantidad = None
columna_codigo = None


for columna in range(
    1,
    hoja.max_column + 1
):

    valor = hoja.cell(
        row=fila_encabezados,
        column=columna
    ).value

    if valor is None:

        continue

    nombre = str(
        valor
    ).strip()

    if nombre == "Comidas":

        columna_producto = columna

    elif nombre == "valor":

        columna_precio = columna

    elif nombre == "Cantidad":

        columna_cantidad = columna

    elif nombre == "Código de serie":

        columna_codigo = columna


# ============================================================
# COMPROBAR COLUMNAS
# ============================================================

if columna_producto is None:

    raise Exception(
        "No se encontró la columna 'Comidas'."
    )


if columna_cantidad is None:

    raise Exception(
        "No se encontró la columna 'Cantidad'."
    )


if columna_codigo is None:

    raise Exception(
        "No se encontró la columna 'Código de serie'."
    )


print(
    f"Columna producto: {columna_producto}"
)

print(
    f"Columna cantidad: {columna_cantidad}"
)

print(
    f"Columna código: {columna_codigo}"
)


if columna_precio is not None:

    print(
        f"Columna precio: {columna_precio}"
    )

else:

    print(
        "Columna precio: no encontrada"
    )

print()


# ============================================================
# CONVERTIR STOCK
# ============================================================

def convertir_cantidad(valor):

    if valor is None:

        return 0

    try:

        return int(
            float(valor)
        )

    except (
        ValueError,
        TypeError
    ):

        return 0


# ============================================================
# BUSCAR PRODUCTO POR CÓDIGO
# ============================================================

def buscar_producto(codigo):

    if codigo is None:

        return None

    codigo = str(
        codigo
    ).strip()

    for fila in range(
        fila_encabezados + 1,
        hoja.max_row + 1
    ):

        valor_codigo = hoja.cell(
            row=fila,
            column=columna_codigo
        ).value

        if valor_codigo is None:

            continue

        codigo_excel = str(
            valor_codigo
        ).strip()

        if codigo_excel == codigo:

            return fila

    return None


# ============================================================
# CREAR / BUSCAR HOJA MOVIMIENTOS
# ============================================================

if nombre_hoja_movimientos in libro.sheetnames:

    hoja_movimientos = libro[
        nombre_hoja_movimientos
    ]

    print(
        "Hoja 'Movimientos': encontrada."
    )

else:

    hoja_movimientos = libro.create_sheet(
        nombre_hoja_movimientos
    )

    print(
        "Hoja 'Movimientos': creada automáticamente."
    )


# ============================================================
# ENCABEZADOS DE MOVIMIENTOS
# ============================================================

encabezados_movimientos = [
    "Fecha",
    "Código",
    "Producto",
    "Operación",
    "Cantidad",
    "Stock anterior",
    "Stock resultante"
]


# ============================================================
# CREAR ENCABEZADOS SI NO EXISTEN
# ============================================================

for columna, nombre in enumerate(
    encabezados_movimientos,
    start=1
):

    celda = hoja_movimientos.cell(
        row=1,
        column=columna
    )

    if celda.value is None:

        celda.value = nombre


# ============================================================
# CREAR BACKUP
# ============================================================

def crear_backup():

    fecha = datetime.now().strftime(
        "%Y-%m-%d_%H-%M-%S"
    )

    nombre_backup = (
        f"productos_{fecha}.xlsx"
    )

    ruta_backup = os.path.join(
        carpeta_backups,
        nombre_backup
    )

    shutil.copy2(
        archivo_excel,
        ruta_backup
    )

    return ruta_backup


# ============================================================
# VALIDAR ARCHIVO EXCEL
# ============================================================

def validar_excel(ruta):

    libro_prueba = None

    try:

        libro_prueba = load_workbook(
            ruta,
            read_only=True
        )

        # Comprobar que exista Movimientos

        if (
            nombre_hoja_movimientos
            not in libro_prueba.sheetnames
        ):

            print()
            print(
                "ERROR: no existe la hoja "
                "'Movimientos' en el archivo guardado."
            )

            return False

        return True

    except Exception as error:

        print()
        print(
            "ERROR: el archivo generado "
            "no pudo ser validado."
        )

        print(
            error
        )

        return False

    finally:

        if libro_prueba is not None:

            try:

                libro_prueba.close()

            except Exception:

                pass


# ============================================================
# GUARDAR EXCEL DE FORMA SEGURA
# ============================================================

def guardar_excel_seguro():

    global libro
    global hoja
    global hoja_movimientos

    directorio = carpeta_programa

    ruta_temporal = None

    try:

        # ----------------------------------------------------
        # CREAR ARCHIVO TEMPORAL
        # ----------------------------------------------------

        archivo_temporal = tempfile.NamedTemporaryFile(
            prefix="stock_temporal_",
            suffix=".xlsx",
            dir=directorio,
            delete=False
        )

        ruta_temporal = archivo_temporal.name

        archivo_temporal.close()

        print()
        print(
            "Guardando Excel..."
        )

        # ----------------------------------------------------
        # GUARDAR
        # ----------------------------------------------------

        libro.save(
            ruta_temporal
        )

        print(
            "Archivo temporal creado."
        )

        # ----------------------------------------------------
        # VALIDAR
        # ----------------------------------------------------

        print(
            "Comprobando archivo..."
        )

        if not validar_excel(
            ruta_temporal
        ):

            print()
            print(
                "El archivo temporal no es válido."
            )

            print(
                "El Excel original NO fue reemplazado."
            )

            return False

        # ----------------------------------------------------
        # REEMPLAZAR ORIGINAL
        # ----------------------------------------------------

        print(
            "Reemplazando archivo original..."
        )

        os.replace(
            ruta_temporal,
            archivo_excel
        )

        ruta_temporal = None

        # ----------------------------------------------------
        # CERRAR LIBRO ACTUAL
        # ----------------------------------------------------

        try:

            libro.close()

        except Exception:

            pass

        # ----------------------------------------------------
        # VOLVER A ABRIR EL EXCEL
        # ----------------------------------------------------

        print(
            "Recargando Excel..."
        )

        libro = load_workbook(
            archivo_excel
        )

        # ----------------------------------------------------
        # VOLVER A BUSCAR HOJA DE PRODUCTOS
        # ----------------------------------------------------

        hoja = buscar_hoja_productos(
            libro
        )

        if hoja is None:

            raise Exception(
                "No se pudo encontrar la hoja "
                "de productos después de guardar."
            )

        # ----------------------------------------------------
        # VOLVER A BUSCAR ENCABEZADOS
        # ----------------------------------------------------

        global fila_encabezados
        global columna_producto
        global columna_precio
        global columna_cantidad
        global columna_codigo

        fila_encabezados = buscar_encabezados(
            hoja
        )

        if fila_encabezados is None:

            raise Exception(
                "No se encontraron los encabezados "
                "después de guardar."
            )

        # ----------------------------------------------------
        # VOLVER A BUSCAR COLUMNAS
        # ----------------------------------------------------

        columna_producto = None
        columna_precio = None
        columna_cantidad = None
        columna_codigo = None

        for columna in range(
            1,
            hoja.max_column + 1
        ):

            valor = hoja.cell(
                row=fila_encabezados,
                column=columna
            ).value

            if valor is None:

                continue

            nombre = str(
                valor
            ).strip()

            if nombre == "Comidas":

                columna_producto = columna

            elif nombre == "valor":

                columna_precio = columna

            elif nombre == "Cantidad":

                columna_cantidad = columna

            elif nombre == "Código de serie":

                columna_codigo = columna

        # ----------------------------------------------------
        # VOLVER A BUSCAR MOVIMIENTOS
        # ----------------------------------------------------

        if (
            nombre_hoja_movimientos
            in libro.sheetnames
        ):

            hoja_movimientos = libro[
                nombre_hoja_movimientos
            ]

        else:

            hoja_movimientos = libro.create_sheet(
                nombre_hoja_movimientos
            )

            for columna, nombre in enumerate(
                encabezados_movimientos,
                start=1
            ):

                hoja_movimientos.cell(
                    row=1,
                    column=columna
                ).value = nombre

        print(
            "Excel guardado y recargado correctamente."
        )

        return True

    except Exception as error:

        print()
        print(
            "=========================================="
        )

        print(
            "       ERROR AL GUARDAR EL EXCEL"
        )

        print(
            "=========================================="
        )

        print()

        print(
            "Mensaje del error:"
        )

        print(
            error
        )

        print()

        print(
            "=========================================="
        )

        print(
            "          TRACEBACK COMPLETO"
        )

        print(
            "=========================================="
        )

        print()

        traceback.print_exc()

        print()

        print(
            "=========================================="
        )

        print(
            "El archivo original NO fue reemplazado."
        )

        print(
            "=========================================="
        )

        print()

        return False

    finally:

        # ----------------------------------------------------
        # ELIMINAR TEMPORAL
        # ----------------------------------------------------

        if (
            ruta_temporal is not None
            and os.path.exists(
                ruta_temporal
            )
        ):

            try:

                os.remove(
                    ruta_temporal
                )

            except Exception:

                pass


# ============================================================
# REGISTRAR MOVIMIENTO
# ============================================================

def registrar_movimiento(
    codigo,
    operacion,
    cantidad
):

    codigo = str(
        codigo
    ).strip()

    operacion = str(
        operacion
    ).strip().upper()

    # --------------------------------------------------------
    # CONVERTIR CANTIDAD
    # --------------------------------------------------------

    try:

        cantidad = int(
            cantidad
        )

    except (
        ValueError,
        TypeError
    ):

        print()
        print(
            "La cantidad debe ser un número entero."
        )

        return

    if cantidad <= 0:

        print()
        print(
            "La cantidad debe ser mayor que 0."
        )

        return

    # --------------------------------------------------------
    # BUSCAR PRODUCTO
    # --------------------------------------------------------

    fila = buscar_producto(
        codigo
    )

    if fila is None:

        print()
        print(
            "Producto no encontrado."
        )

        print(
            f"Código: {codigo}"
        )

        return

    # --------------------------------------------------------
    # OBTENER PRODUCTO
    # --------------------------------------------------------

    producto = hoja.cell(
        row=fila,
        column=columna_producto
    ).value

    # --------------------------------------------------------
    # CELDA DEL STOCK
    # --------------------------------------------------------

    celda_stock = hoja.cell(
        row=fila,
        column=columna_cantidad
    )

    stock_anterior = convertir_cantidad(
        celda_stock.value
    )

    # ========================================================
    # CALCULAR STOCK NUEVO
    # ========================================================

    if operacion == "ENTRADA":

        stock_nuevo = (
            stock_anterior
            + cantidad
        )

    elif operacion == "SALIDA":

        if cantidad > stock_anterior:

            print()
            print(
                "=========================================="
            )

            print(
                "          STOCK INSUFICIENTE"
            )

            print(
                "=========================================="
            )

            print()

            print(
                f"Producto: {producto}"
            )

            print(
                f"Stock disponible: {stock_anterior}"
            )

            print(
                f"Cantidad solicitada: {cantidad}"
            )

            print()

            return

        stock_nuevo = (
            stock_anterior
            - cantidad
        )

    else:

        print()
        print(
            "Operación inválida."
        )

        return

    # ========================================================
    # CREAR BACKUP
    # ========================================================

    try:

        ruta_backup = crear_backup()

    except Exception as error:

        print()
        print(
            "=========================================="
        )

        print(
            "       NO SE PUDO CREAR EL BACKUP"
        )

        print(
            "=========================================="
        )

        print()

        print(
            error
        )

        print()

        print(
            "El Excel no fue modificado."
        )

        return

    # ========================================================
    # GUARDAR VALOR ORIGINAL
    # ========================================================

    valor_stock_original = celda_stock.value

    # ========================================================
    # MODIFICAR STOCK
    # ========================================================

    celda_stock.value = stock_nuevo

    # ========================================================
    # CREAR REGISTRO DE MOVIMIENTO
    # ========================================================

    nueva_fila = (
        hoja_movimientos.max_row + 1
    )

    hoja_movimientos.cell(
        row=nueva_fila,
        column=1
    ).value = datetime.now()

    hoja_movimientos.cell(
        row=nueva_fila,
        column=2
    ).value = codigo

    hoja_movimientos.cell(
        row=nueva_fila,
        column=3
    ).value = producto

    hoja_movimientos.cell(
        row=nueva_fila,
        column=4
    ).value = operacion

    hoja_movimientos.cell(
        row=nueva_fila,
        column=5
    ).value = cantidad

    hoja_movimientos.cell(
        row=nueva_fila,
        column=6
    ).value = stock_anterior

    hoja_movimientos.cell(
        row=nueva_fila,
        column=7
    ).value = stock_nuevo

    # ========================================================
    # GUARDAR
    # ========================================================

    guardado = guardar_excel_seguro()

    # ========================================================
    # SI FALLÓ
    # ========================================================

    if not guardado:

        print()
        print(
            "=========================================="
        )

        print(
            "       CAMBIO NO APLICADO"
        )

        print(
            "=========================================="
        )

        print()

        print(
            "El archivo original fue conservado."
        )

        print(
            "Backup disponible en:"
        )

        print(
            ruta_backup
        )

        print()

        return

    # ========================================================
    # RESULTADO
    # ========================================================

    print()
    print(
        "=========================================="
    )

    print(
        "          STOCK ACTUALIZADO"
    )

    print(
        "=========================================="
    )

    print()

    print(
        f"Producto: {producto}"
    )

    print(
        f"Código: {codigo}"
    )

    print(
        f"Operación: {operacion}"
    )

    print(
        f"Cantidad: {cantidad}"
    )

    print(
        f"Stock anterior: {stock_anterior}"
    )

    print(
        f"Stock actual: {stock_nuevo}"
    )

    print()

    print(
        f"Backup creado: {ruta_backup}"
    )

    print()


# ============================================================
# MOSTRAR PRODUCTO
# ============================================================

def mostrar_producto(codigo):

    fila = buscar_producto(
        codigo
    )

    if fila is None:

        print()
        print(
            "Producto no encontrado."
        )

        return

    producto = hoja.cell(
        row=fila,
        column=columna_producto
    ).value

    stock = convertir_cantidad(
        hoja.cell(
            row=fila,
            column=columna_cantidad
        ).value
    )

    codigo_excel = hoja.cell(
        row=fila,
        column=columna_codigo
    ).value

    precio = None

    if columna_precio is not None:

        precio = hoja.cell(
            row=fila,
            column=columna_precio
        ).value

    print()
    print(
        "=========================================="
    )

    print(
        "              PRODUCTO"
    )

    print(
        "=========================================="
    )

    print()

    print(
        f"Producto: {producto}"
    )

    print(
        f"Código: {codigo_excel}"
    )

    print(
        f"Precio: {precio}"
    )

    print(
        f"Stock: {stock}"
    )

    if stock <= stock_minimo:

        print()

        print(
            "⚠ ADVERTENCIA: STOCK BAJO"
        )

    print()


# ============================================================
# MOSTRAR STOCK BAJO
# ============================================================

def mostrar_stock_bajo():

    encontrados = 0

    print()
    print(
        "=========================================="
    )

    print(
        "             STOCK BAJO"
    )

    print(
        "=========================================="
    )

    print()

    for fila in range(
        fila_encabezados + 1,
        hoja.max_row + 1
    ):

        producto = hoja.cell(
            row=fila,
            column=columna_producto
        ).value

        codigo = hoja.cell(
            row=fila,
            column=columna_codigo
        ).value

        if (
            producto is None
            or codigo is None
        ):

            continue

        stock = convertir_cantidad(
            hoja.cell(
                row=fila,
                column=columna_cantidad
            ).value
        )

        if stock <= stock_minimo:

            print(
                f"{producto} | "
                f"Código: {codigo} | "
                f"Stock: {stock}"
            )

            encontrados += 1

    if encontrados == 0:

        print(
            "No hay productos con stock bajo."
        )

    print()


# ============================================================
# MOSTRAR MOVIMIENTOS
# ============================================================

def mostrar_movimientos():

    print()
    print(
        "=========================================="
    )

    print(
        "             MOVIMIENTOS"
    )

    print(
        "=========================================="
    )

    print()

    if hoja_movimientos.max_row <= 1:

        print(
            "Todavía no hay movimientos."
        )

        print()

        return

    for fila in range(
        2,
        hoja_movimientos.max_row + 1
    ):

        fecha = hoja_movimientos.cell(
            row=fila,
            column=1
        ).value

        codigo = hoja_movimientos.cell(
            row=fila,
            column=2
        ).value

        producto = hoja_movimientos.cell(
            row=fila,
            column=3
        ).value

        operacion = hoja_movimientos.cell(
            row=fila,
            column=4
        ).value

        cantidad = hoja_movimientos.cell(
            row=fila,
            column=5
        ).value

        stock = hoja_movimientos.cell(
            row=fila,
            column=7
        ).value

        print(
            f"{fecha} | "
            f"{codigo} | "
            f"{producto} | "
            f"{operacion} | "
            f"{cantidad} | "
            f"Stock: {stock}"
        )

    print()


# ============================================================
# MENÚ PRINCIPAL
# ============================================================

while True:

    print()
    print(
        "=========================================="
    )

    print(
        "          SISTEMA DE INVENTARIO"
    )

    print(
        "=========================================="
    )

    print()

    print(
        "1 - Entrada de stock"
    )

    print(
        "2 - Salida de stock"
    )

    print(
        "3 - Buscar producto"
    )

    print(
        "4 - Ver stock bajo"
    )

    print(
        "5 - Ver movimientos"
    )

    print(
        "6 - Salir"
    )

    print()

    opcion = input(
        "Seleccione una opción: "
    ).strip()

    # ========================================================
    # ENTRADA
    # ========================================================

    if opcion == "1":

        print()

        codigo = input(
            "Código del producto: "
        ).strip()

        cantidad = input(
            "Cantidad que ingresa: "
        ).strip()

        registrar_movimiento(
            codigo,
            "ENTRADA",
            cantidad
        )

    # ========================================================
    # SALIDA
    # ========================================================

    elif opcion == "2":

        print()

        codigo = input(
            "Código del producto: "
        ).strip()

        cantidad = input(
            "Cantidad que sale: "
        ).strip()

        registrar_movimiento(
            codigo,
            "SALIDA",
            cantidad
        )

    # ========================================================
    # BUSCAR PRODUCTO
    # ========================================================

    elif opcion == "3":

        codigo = input(
            "Código del producto: "
        ).strip()

        mostrar_producto(
            codigo
        )

    # ========================================================
    # STOCK BAJO
    # ========================================================

    elif opcion == "4":

        mostrar_stock_bajo()

    # ========================================================
    # MOVIMIENTOS
    # ========================================================

    elif opcion == "5":

        mostrar_movimientos()

    # ========================================================
    # SALIR
    # ========================================================

    elif opcion == "6":

        print()
        print(
            "Sistema cerrado."
        )

        try:

            libro.close()

        except Exception:

            pass

        break

    # ========================================================
    # OPCIÓN INCORRECTA
    # ========================================================

    else:

        print()
        print(
            "Opción no válida."
        )