import os
from openpyxl import load_workbook


# ============================================================
# CONFIGURACIÓN
# ============================================================

archivo_original = "productos_con_barras.xlsx"

archivo_corregido = "productos_corregido.xlsx"

carpeta_codigos = "codigos"


# ============================================================
# ABRIR EXCEL
# ============================================================

libro = load_workbook(
    archivo_original
)

hoja = libro.active


# ============================================================
# BUSCAR FILA DE ENCABEZADOS
# ============================================================

fila_encabezados = None

for fila in range(
    1,
    hoja.max_row + 1
):

    valores = []

    for columna in range(
        1,
        hoja.max_column + 1
    ):

        valor = hoja.cell(
            row=fila,
            column=columna
        ).value

        if valor is not None:

            valores.append(
                str(valor).strip()
            )


    if (
        "Comidas" in valores
        and "Código de serie" in valores
    ):

        fila_encabezados = fila

        break


if fila_encabezados is None:

    raise Exception(
        "No se encontró la fila de encabezados."
    )


print(
    "Fila de encabezados:",
    fila_encabezados
)


# ============================================================
# BUSCAR COLUMNAS
# ============================================================

columna_producto = None

columna_precio = None

columna_codigo = None


for columna in range(
    1,
    hoja.max_column + 1
):

    valor = hoja.cell(
        row=fila_encabezados,
        column=columna
    ).value


    if valor is None:
        continue


    nombre = str(
        valor
    ).strip()


    if nombre == "Comidas":

        columna_producto = columna


    elif nombre == "valor":

        columna_precio = columna


    elif nombre == "Código de serie":

        columna_codigo = columna


print(
    "Columna producto:",
    columna_producto
)

print(
    "Columna precio:",
    columna_precio
)

print(
    "Columna código:",
    columna_codigo
)

print()


# ============================================================
# COMPROBAR COLUMNAS
# ============================================================

if columna_producto is None:

    raise Exception(
        "No se encontró la columna 'Comidas'."
    )


if columna_precio is None:

    raise Exception(
        "No se encontró la columna 'valor'."
    )


if columna_codigo is None:

    raise Exception(
        "No se encontró la columna 'Código de serie'."
    )


# ============================================================
# BUSCAR CATEGORÍAS
# ============================================================

categorias = []


for fila in range(
    fila_encabezados + 1,
    hoja.max_row + 1
):

    producto = hoja.cell(
        row=fila,
        column=columna_producto
    ).value


    precio = hoja.cell(
        row=fila,
        column=columna_precio
    ).value


    if producto is None:
        continue


    producto = str(
        producto
    ).strip()


    if producto == "":
        continue


    # --------------------------------------------------------
    # SIN PRECIO = CATEGORÍA
    # --------------------------------------------------------

    if precio is None:

        categorias.append(
            fila
        )


# ============================================================
# MOSTRAR CATEGORÍAS ENCONTRADAS
# ============================================================

print(
    "Categorías encontradas:"
)

for fila in categorias:

    nombre = hoja.cell(
        row=fila,
        column=columna_producto
    ).value

    codigo = hoja.cell(
        row=fila,
        column=columna_codigo
    ).value

    print(
        f"{nombre} -> código actual: {codigo}"
    )


print()


# ============================================================
# ELIMINAR CÓDIGOS DE LAS CATEGORÍAS
# ============================================================

for fila in categorias:

    celda_codigo = hoja.cell(
        row=fila,
        column=columna_codigo
    )


    codigo = celda_codigo.value


    if codigo is None:
        continue


    codigo = str(
        codigo
    ).strip()


    # --------------------------------------------------------
    # ELIMINAR CÓDIGO DEL EXCEL
    # --------------------------------------------------------

    celda_codigo.value = None


    print(
        f"Código eliminado de categoría: {codigo}"
    )


    # --------------------------------------------------------
    # BUSCAR IMAGEN
    # --------------------------------------------------------

    ruta_imagen = os.path.join(
        carpeta_codigos,
        f"{codigo}.png"
    )


    # --------------------------------------------------------
    # ELIMINAR IMAGEN
    # --------------------------------------------------------

    if os.path.exists(
        ruta_imagen
    ):

        os.remove(
            ruta_imagen
        )

        print(
            f"Imagen eliminada: {codigo}.png"
        )


# ============================================================
# ELIMINAR IMÁGENES DE LAS CATEGORÍAS DEL EXCEL
# ============================================================

imagenes_a_eliminar = []


for imagen in hoja._images:

    # --------------------------------------------------------
    # OBTENER FILA DONDE ESTÁ LA IMAGEN
    # --------------------------------------------------------

    fila_imagen = (
        imagen.anchor._from.row + 1
    )


    if fila_imagen in categorias:

        imagenes_a_eliminar.append(
            imagen
        )


for imagen in imagenes_a_eliminar:

    hoja._images.remove(
        imagen
    )


# ============================================================
# GUARDAR
# ============================================================

libro.save(
    archivo_corregido
)


# ============================================================
# FINAL
# ============================================================

print()

print(
    "=========================================="
)

print(
    "       CORRECCIÓN TERMINADA"
)

print(
    "=========================================="
)

print()

print(
    "Archivo original:",
    archivo_original
)

print(
    "Archivo corregido:",
    archivo_corregido
)

print()

print(
    "Las categorías quedaron sin código."
)

print(
    "Los códigos de los productos NO fueron modificados."
)

print()