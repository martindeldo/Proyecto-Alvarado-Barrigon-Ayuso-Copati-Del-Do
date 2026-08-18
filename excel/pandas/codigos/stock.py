import os
import shutil
from datetime import datetime

from openpyxl import load_workbook


# ============================================================
# CONFIGURACIÓN
# ============================================================

archivo_excel = "productos_con_barras.xlsx"

carpeta_backups = "backups"

nombre_hoja_movimientos = "Movimientos"

stock_minimo = 5


# ============================================================
# CREAR CARPETA DE BACKUPS
# ============================================================

os.makedirs(
    carpeta_backups,
    exist_ok=True
)


# ============================================================
# BUSCAR FILA DE ENCABEZADOS
# ============================================================

def buscar_encabezados(hoja):

    for fila in range(
        1,
        hoja.max_row + 1
    ):

        valores = []

        for columna in range(
            1,
            hoja.max_column + 1
        ):

            valor = hoja.cell(
                row=fila,
                column=columna
            ).value


            if valor is not None:

                valores.append(
                    str(valor).strip()
                )


        if (
            "Comidas" in valores
            and "Código de serie" in valores
        ):

            return fila


    return None


# ============================================================
# ABRIR EXCEL
# ============================================================

if not os.path.exists(
    archivo_excel
):

    raise Exception(
        f"No existe el archivo: "
        f"{archivo_excel}"
    )


libro = load_workbook(
    archivo_excel
)

hoja = libro.active


# ============================================================
# BUSCAR ENCABEZADOS
# ============================================================

fila_encabezados = buscar_encabezados(
    hoja
)


if fila_encabezados is None:

    raise Exception(
        "No se encontró la fila de encabezados."
    )


print()
print(
    f"Fila de encabezados: "
    f"{fila_encabezados}"
)


# ============================================================
# BUSCAR COLUMNAS
# ============================================================

columna_producto = None
columna_precio = None
columna_cantidad = None
columna_codigo = None


for columna in range(
    1,
    hoja.max_column + 1
):

    valor = hoja.cell(
        row=fila_encabezados,
        column=columna
    ).value


    if valor is None:

        continue


    nombre = str(
        valor
    ).strip()


    if nombre == "Comidas":

        columna_producto = columna


    elif nombre == "valor":

        columna_precio = columna


    elif nombre == "Cantidad":

        columna_cantidad = columna


    elif nombre == "Código de serie":

        columna_codigo = columna


# ============================================================
# COMPROBAR
# ============================================================

if columna_producto is None:

    raise Exception(
        "No se encontró 'Comidas'."
    )


if columna_cantidad is None:

    raise Exception(
        "No se encontró 'Cantidad'."
    )


if columna_codigo is None:

    raise Exception(
        "No se encontró 'Código de serie'."
    )


print(
    f"Columna producto: "
    f"{columna_producto}"
)

print(
    f"Columna cantidad: "
    f"{columna_cantidad}"
)

print(
    f"Columna código: "
    f"{columna_codigo}"
)

print()


# ============================================================
# CONVERTIR STOCK
# ============================================================

def convertir_cantidad(valor):

    if valor is None:

        return 0


    try:

        return int(
            float(valor)
        )


    except (
        ValueError,
        TypeError
    ):

        return 0


# ============================================================
# BUSCAR PRODUCTO POR CÓDIGO
# ============================================================

def buscar_producto(codigo):

    codigo = str(
        codigo
    ).strip()


    for fila in range(
        fila_encabezados + 1,
        hoja.max_row + 1
    ):

        valor_codigo = hoja.cell(
            row=fila,
            column=columna_codigo
        ).value


        if valor_codigo is None:

            continue


        codigo_excel = str(
            valor_codigo
        ).strip()


        if codigo_excel == codigo:

            return fila


    return None


# ============================================================
# CREAR HOJA MOVIMIENTOS
# ============================================================

if nombre_hoja_movimientos in libro.sheetnames:

    hoja_movimientos = libro[
        nombre_hoja_movimientos
    ]

else:

    hoja_movimientos = libro.create_sheet(
        nombre_hoja_movimientos
    )


# ============================================================
# ENCABEZADOS MOVIMIENTOS
# ============================================================

encabezados = [
    "Fecha",
    "Código",
    "Producto",
    "Operación",
    "Cantidad",
    "Stock anterior",
    "Stock resultante"
]


if (
    hoja_movimientos.max_row == 1
    and hoja_movimientos.cell(
        row=1,
        column=1
    ).value is None
):

    for columna, nombre in enumerate(
        encabezados,
        start=1
    ):

        hoja_movimientos.cell(
            row=1,
            column=columna
        ).value = nombre


# ============================================================
# BACKUP
# ============================================================

def crear_backup():

    fecha = datetime.now().strftime(
        "%Y-%m-%d_%H-%M-%S"
    )


    nombre_backup = (
        f"productos_{fecha}.xlsx"
    )


    ruta_backup = os.path.join(
        carpeta_backups,
        nombre_backup
    )


    shutil.copy2(
        archivo_excel,
        ruta_backup
    )


    return ruta_backup


# ============================================================
# REGISTRAR MOVIMIENTO
# ============================================================

def registrar_movimiento(
    codigo,
    operacion,
    cantidad
):

    codigo = str(
        codigo
    ).strip()


    try:

        cantidad = int(
            cantidad
        )


    except ValueError:

        print(
            "La cantidad debe ser un número entero."
        )

        return


    if cantidad <= 0:

        print(
            "La cantidad debe ser mayor que 0."
        )

        return


    # --------------------------------------------------------
    # BUSCAR PRODUCTO
    # --------------------------------------------------------

    fila = buscar_producto(
        codigo
    )


    if fila is None:

        print()
        print(
            "Producto no encontrado."
        )
        print(
            f"Código: {codigo}"
        )

        return


    # --------------------------------------------------------
    # PRODUCTO
    # --------------------------------------------------------

    producto = hoja.cell(
        row=fila,
        column=columna_producto
    ).value


    # --------------------------------------------------------
    # STOCK ACTUAL
    # --------------------------------------------------------

    celda_stock = hoja.cell(
        row=fila,
        column=columna_cantidad
    )


    stock_anterior = convertir_cantidad(
        celda_stock.value
    )


    # ========================================================
    # ENTRADA
    # ========================================================

    if operacion == "ENTRADA":

        stock_nuevo = (
            stock_anterior
            + cantidad
        )


    # ========================================================
    # SALIDA
    # ========================================================

    elif operacion == "SALIDA":

        if cantidad > stock_anterior:

            print()

            print(
                "=========================================="
            )

            print(
                "          STOCK INSUFICIENTE"
            )

            print(
                "=========================================="
            )

            print()

            print(
                f"Producto: {producto}"
            )

            print(
                f"Stock disponible: "
                f"{stock_anterior}"
            )

            print(
                f"Cantidad solicitada: "
                f"{cantidad}"
            )

            print()

            return


        stock_nuevo = (
            stock_anterior
            - cantidad
        )


    else:

        print(
            "Operación inválida."
        )

        return


    # ========================================================
    # CREAR BACKUP
    # ========================================================

    try:

        ruta_backup = crear_backup()


    except Exception as error:

        print()

        print(
            "No se pudo crear el backup."
        )

        print(
            error
        )

        return


    # ========================================================
    # ACTUALIZAR STOCK
    # ========================================================

    celda_stock.value = (
        stock_nuevo
    )


    # ========================================================
    # REGISTRAR MOVIMIENTO
    # ========================================================

    nueva_fila = (
        hoja_movimientos.max_row + 1
    )


    hoja_movimientos.cell(
        row=nueva_fila,
        column=1
    ).value = datetime.now()


    hoja_movimientos.cell(
        row=nueva_fila,
        column=2
    ).value = codigo


    hoja_movimientos.cell(
        row=nueva_fila,
        column=3
    ).value = producto


    hoja_movimientos.cell(
        row=nueva_fila,
        column=4
    ).value = operacion


    hoja_movimientos.cell(
        row=nueva_fila,
        column=5
    ).value = cantidad


    hoja_movimientos.cell(
        row=nueva_fila,
        column=6
    ).value = stock_anterior


    hoja_movimientos.cell(
        row=nueva_fila,
        column=7
    ).value = stock_nuevo


    # ========================================================
    # GUARDAR
    # ========================================================

    try:

        libro.save(
            archivo_excel
        )


    except Exception as error:

        print()

        print(
            "ERROR AL GUARDAR EL EXCEL."
        )

        print(
            error
        )

        print()

        print(
            "El backup está en:"
        )

        print(
            ruta_backup
        )

        return


    # ========================================================
    # RESULTADO
    # ========================================================

    print()

    print(
        "=========================================="
    )

    print(
        "          STOCK ACTUALIZADO"
    )

    print(
        "=========================================="
    )

    print()

    print(
        f"Producto: {producto}"
    )

    print(
        f"Código: {codigo}"
    )

    print(
        f"Operación: {operacion}"
    )

    print(
        f"Cantidad: {cantidad}"
    )

    print(
        f"Stock anterior: {stock_anterior}"
    )

    print(
        f"Stock actual: {stock_nuevo}"
    )

    print()

    print(
        f"Backup creado: {ruta_backup}"
    )


# ============================================================
# BUSCAR PRODUCTO
# ============================================================

def mostrar_producto(codigo):

    fila = buscar_producto(
        codigo
    )


    if fila is None:

        print()
        print(
            "Producto no encontrado."
        )

        return


    producto = hoja.cell(
        row=fila,
        column=columna_producto
    ).value


    stock = convertir_cantidad(
        hoja.cell(
            row=fila,
            column=columna_cantidad
        ).value
    )


    codigo_excel = hoja.cell(
        row=fila,
        column=columna_codigo
    ).value


    precio = None


    if columna_precio is not None:

        precio = hoja.cell(
            row=fila,
            column=columna_precio
        ).value


    print()

    print(
        "=========================================="
    )

    print(
        "              PRODUCTO"
    )

    print(
        "=========================================="
    )

    print()

    print(
        f"Producto: {producto}"
    )

    print(
        f"Código: {codigo_excel}"
    )

    print(
        f"Precio: {precio}"
    )

    print(
        f"Stock: {stock}"
    )

    if stock <= stock_minimo:

        print()

        print(
            "⚠ ADVERTENCIA: STOCK BAJO"
        )

    print()


# ============================================================
# MOSTRAR STOCK BAJO
# ============================================================

def mostrar_stock_bajo():

    encontrados = 0


    print()

    print(
        "=========================================="
    )

    print(
        "             STOCK BAJO"
    )

    print(
        "=========================================="
    )

    print()


    for fila in range(
        fila_encabezados + 1,
        hoja.max_row + 1
    ):

        producto = hoja.cell(
            row=fila,
            column=columna_producto
        ).value


        codigo = hoja.cell(
            row=fila,
            column=columna_codigo
        ).value


        if producto is None or codigo is None:

            continue


        stock = convertir_cantidad(
            hoja.cell(
                row=fila,
                column=columna_cantidad
            ).value
        )


        if stock <= stock_minimo:

            print(
                f"{producto} | "
                f"Código: {codigo} | "
                f"Stock: {stock}"
            )

            encontrados += 1


    if encontrados == 0:

        print(
            "No hay productos con stock bajo."
        )


    print()


# ============================================================
# MOSTRAR MOVIMIENTOS
# ============================================================

def mostrar_movimientos():

    print()

    print(
        "=========================================="
    )

    print(
        "             MOVIMIENTOS"
    )

    print(
        "=========================================="
    )

    print()


    if hoja_movimientos.max_row <= 1:

        print(
            "Todavía no hay movimientos."
        )

        return


    for fila in range(
        2,
        hoja_movimientos.max_row + 1
    ):

        fecha = hoja_movimientos.cell(
            row=fila,
            column=1
        ).value


        codigo = hoja_movimientos.cell(
            row=fila,
            column=2
        ).value


        producto = hoja_movimientos.cell(
            row=fila,
            column=3
        ).value


        operacion = hoja_movimientos.cell(
            row=fila,
            column=4
        ).value


        cantidad = hoja_movimientos.cell(
            row=fila,
            column=5
        ).value


        stock = hoja_movimientos.cell(
            row=fila,
            column=7
        ).value


        print(
            f"{fecha} | "
            f"{codigo} | "
            f"{producto} | "
            f"{operacion} | "
            f"{cantidad} | "
            f"Stock: {stock}"
        )


    print()


# ============================================================
# GUARDAR CAMBIOS INICIALES
# ============================================================

libro.save(
    archivo_excel
)


# ============================================================
# MENÚ PRINCIPAL
# ============================================================

while True:

    print()

    print(
        "=========================================="
    )

    print(
        "          SISTEMA DE INVENTARIO"
    )

    print(
        "=========================================="
    )

    print()

    print(
        "1 - Entrada de stock"
    )

    print(
        "2 - Salida de stock"
    )

    print(
        "3 - Buscar producto"
    )

    print(
        "4 - Ver stock bajo"
    )

    print(
        "5 - Ver movimientos"
    )

    print(
        "6 - Salir"
    )

    print()


    opcion = input(
        "Seleccione una opción: "
    ).strip()


    # ========================================================
    # ENTRADA
    # ========================================================

    if opcion == "1":

        print()

        codigo = input(
            "Código del producto: "
        ).strip()


        cantidad = input(
            "Cantidad que ingresa: "
        ).strip()


        registrar_movimiento(
            codigo,
            "ENTRADA",
            cantidad
        )


    # ========================================================
    # SALIDA
    # ========================================================

    elif opcion == "2":

        print()

        codigo = input(
            "Código del producto: "
        ).strip()


        cantidad = input(
            "Cantidad que sale: "
        ).strip()


        registrar_movimiento(
            codigo,
            "SALIDA",
            cantidad
        )


    # ========================================================
    # BUSCAR
    # ========================================================

    elif opcion == "3":

        codigo = input(
            "Código del producto: "
        ).strip()


        mostrar_producto(
            codigo
        )


    # ========================================================
    # STOCK BAJO
    # ========================================================

    elif opcion == "4":

        mostrar_stock_bajo()


    # ========================================================
    # MOVIMIENTOS
    # ========================================================

    elif opcion == "5":

        mostrar_movimientos()


    # ========================================================
    # SALIR
    # ========================================================

    elif opcion == "6":

        print()

        print(
            "Sistema cerrado."
        )

        break


    else:

        print()

        print(
            "Opción no válida."
        )