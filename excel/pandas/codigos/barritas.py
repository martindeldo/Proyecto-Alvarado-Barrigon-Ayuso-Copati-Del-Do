import os
import pandas as pd
import barcode
from barcode.writer import ImageWriter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image


# ============================================================
# CONFIGURACIÓN
# ============================================================

archivo_excel = "productos.xlsx"
carpeta_codigos = "codigos"

codigo_inicial = 100000000001


# ============================================================
# CREAR CARPETA DE CÓDIGOS
# ============================================================

os.makedirs(carpeta_codigos, exist_ok=True)


# ============================================================
# LEER EXCEL CON PANDAS
# ============================================================

df = pd.read_excel(
    archivo_excel,
    header=1,
    dtype={"Código de serie": "string"}
)

print("Columnas encontradas por Pandas:")
print(df.columns)
print()


# ============================================================
# ABRIR EXCEL CON OPENPYXL
# ============================================================

libro = load_workbook(archivo_excel)

hoja = libro.active


# ============================================================
# BUSCAR AUTOMÁTICAMENTE LA FILA DE ENCABEZADOS
# ============================================================

fila_encabezados = None

for fila in range(1, hoja.max_row + 1):

    valores = []

    for columna in range(1, hoja.max_column + 1):

        valor = hoja.cell(
            row=fila,
            column=columna
        ).value

        if valor is not None:
            valores.append(str(valor).strip())

    # Mostrar las primeras filas para comprobar
    if fila <= 5:
        print(
            f"Fila {fila}: {valores}"
        )

    # Buscar nuestros encabezados
    if (
        "Comidas" in valores
        and "Código de serie" in valores
    ):
        fila_encabezados = fila
        break


# ============================================================
# COMPROBAR FILA DE ENCABEZADOS
# ============================================================

if fila_encabezados is None:

    raise Exception(
        "No se encontró la fila de encabezados."
    )


print()
print(
    f"Fila de encabezados encontrada: {fila_encabezados}"
)
print()


# ============================================================
# BUSCAR COLUMNAS
# ============================================================

columna_producto = None
columna_precio = None
columna_cantidad = None
columna_codigo = None


for columna in range(1, hoja.max_column + 1):

    valor = hoja.cell(
        row=fila_encabezados,
        column=columna
    ).value

    if valor is None:
        continue

    nombre = str(valor).strip()

    if nombre == "Comidas":

        columna_producto = columna

    elif nombre == "valor":

        columna_precio = columna

    elif nombre == "Cantidad":

        columna_cantidad = columna

    elif nombre == "Código de serie":

        columna_codigo = columna


# ============================================================
# MOSTRAR COLUMNAS
# ============================================================

print("Columnas detectadas:")

print(
    "Producto:",
    columna_producto
)

print(
    "Precio:",
    columna_precio
)

print(
    "Cantidad:",
    columna_cantidad
)

print(
    "Código:",
    columna_codigo
)

print()


# ============================================================
# COMPROBAR COLUMNAS IMPORTANTES
# ============================================================

if columna_producto is None:

    raise Exception(
        "No se encontró la columna 'Comidas'."
    )


if columna_codigo is None:

    raise Exception(
        "No se encontró la columna 'Código de serie'."
    )


# ============================================================
# BUSCAR EL ÚLTIMO CÓDIGO EXISTENTE
# ============================================================

ultimo_codigo = codigo_inicial - 1


for fila in range(
    fila_encabezados + 1,
    hoja.max_row + 1
):

    valor_codigo = hoja.cell(
        row=fila,
        column=columna_codigo
    ).value

    if valor_codigo is None:
        continue

    try:

        numero = int(
            float(valor_codigo)
        )

        if numero > ultimo_codigo:

            ultimo_codigo = numero

    except (ValueError, TypeError):

        continue


print(
    "Último código encontrado:",
    ultimo_codigo
)

print()


# ============================================================
# PROCESAR PRODUCTOS
# ============================================================

for fila in range(
    fila_encabezados + 1,
    hoja.max_row + 1
):

    # --------------------------------------------------------
    # OBTENER PRODUCTO
    # --------------------------------------------------------

    producto = hoja.cell(
        row=fila,
        column=columna_producto
    ).value

    if producto is None:
        continue

    producto = str(producto).strip()

    if producto == "":
        continue


    # --------------------------------------------------------
    # OBTENER PRECIO
    # --------------------------------------------------------

    precio = None

    if columna_precio is not None:

        precio = hoja.cell(
            row=fila,
            column=columna_precio
        ).value


    # --------------------------------------------------------
    # CELDA DEL CÓDIGO
    # --------------------------------------------------------

    celda_codigo = hoja.cell(
        row=fila,
        column=columna_codigo
    )

    codigo_existente = celda_codigo.value


    # ========================================================
    # PRODUCTO CON CÓDIGO
    # ========================================================

    if (
        codigo_existente is not None
        and str(codigo_existente).strip() != ""
    ):

        codigo = str(
            codigo_existente
        ).strip()

        print(
            f"{producto} -> código existente: {codigo}"
        )


    # ========================================================
    # PRODUCTO SIN CÓDIGO
    # ========================================================

    else:

        ultimo_codigo += 1

        codigo = str(
            ultimo_codigo
        )

        celda_codigo.value = codigo

        print(
            f"{producto} -> NUEVO código: {codigo}"
        )


    # ========================================================
    # RUTA DEL CÓDIGO DE BARRAS
    # ========================================================

    ruta_imagen = os.path.join(
        carpeta_codigos,
        f"{codigo}.png"
    )


    # ========================================================
    # CREAR CÓDIGO DE BARRAS
    # ========================================================

    if not os.path.exists(ruta_imagen):

        codigo_barras = barcode.get(
            "code128",
            codigo,
            writer=ImageWriter()
        )

        codigo_barras.save(
            os.path.join(
                carpeta_codigos,
                codigo
            )
        )

        print(
            f"    Código de barras creado: {codigo}.png"
        )


    # ========================================================
    # INSERTAR IMAGEN
    # ========================================================

    if os.path.exists(ruta_imagen):

        imagen = Image(
            ruta_imagen
        )

        imagen.width = 180
        imagen.height = 60

        hoja.add_image(
            imagen,
            celda_codigo.coordinate
        )

        hoja.row_dimensions[
            fila
        ].height = 50


# ============================================================
# GUARDAR EXCEL
# ============================================================

libro.save(
    archivo_excel
)


# ============================================================
# FINAL
# ============================================================

print()
print("==========================================")
print("       INVENTARIO ACTUALIZADO")
print("==========================================")
print()
print(
    f"Excel: {archivo_excel}"
)
print(
    f"Códigos: {carpeta_codigos}"
)
print()
print(
    "Los códigos existentes fueron conservados."
)
print(
    "Los productos nuevos recibieron códigos nuevos."
)
print()